# -*- coding: utf-8 -*-
"""
Created on Thu Sep 24 00:43:26 2020

@author: golge
"""

import time
print("ENTER'a basarsan başlar, CNTRL+C'ye basarsan durur")
while True:
    try:
        input()
        start_time = time.time()
        print("Başladı")
    except KeyboardInterrupt:
        print("Durdu")
        end_time = time.time()
        print("Hepi-topu ",round(end_time - start_time, 2)," saniye geçmiş...")
        break
#%% Yazıyı konuşmaya çevirme şeysi

from gtts import gTTS
from playsound import playsound

audio = 'speech.mp3'
language = 'en'
sp =gTTS(text ="Seslendirilmesini istediğiniz yazıyı yazınız:  ", lang = language,slow=False)

sp.save(audio)
playsound(audio)

#%% Text çevirmeni

from translate import Translator

translator =Translator(from_lang = 'Spanish', to_lang='english')
result = translator.translate('Te amo')
print(result)

#%% print(value,sep,end,file,flush)
print("a","b","c")
print("a" "b" "c")
print("a","b","c",sep="-")
print("a","b","c",end="*")
print("a","b","c",fie=file_obj)
file_obj.close()

#%% Yazım kılavuzu

from textblob import textblob

text = TextBlob("he is my favrote hero")
print(text.correct())

#%% URL kısaltıcı

import pyshorteners

link = 'https://www.instagram.com/p/CD5pxxAgr2n/'

print(pyshorteners.Shortener().clckru.short(link))

#%% Whatsapp mesajı atan uygulama

import pywhatkit as kit

kit.sendwhatmsg("+905*******","Selam kardeş",11,32)

#%% sistem ekipmanlarını ğrenme kafası

import platform
my_system = platform.uname()
print(f"System: {my_system.system}")
print(f"Node Name: {my_system.node}")
print(f"Release: {my_system.release}")
print(f"Version: {my_system.version}")
print(f"Machine: {my_system.machine}")
print(f"Processor: {my_system.processor}")

#%% Değerlerin kapladığı alan

import sys
a,b,c,d = "abcde","xy",2,15.06
print(sys.getsizeof(a))
print(sys.getsizeof(b))
print(sys.getsizeof(c))
print(sys.getsizeof(d))

#%% Rakamı ters çevirme mat'ı

number = 12345
remainder = []
while number !=0:
    remainder.append(number%10)
    number //=10

print("".join(map(str,remainder)))

#%% BİLGİSAYARI KAPATIR VALLAA

import os
os.system("shutdown /s /t l")

#%% Oto-Yazar(mış gibilerinden)

import pyautogui

pyautogui.typewrite('Hello, World!')

pyautogui.typewrite('Hello, World',interval=0.25)

#%% Excell

from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws['A1'] = 42

ws.append([1,2,3])

import datetime
ws['A2'] = datetime.datetime.now()

wb.save("sample.xlsx")

#%% Fraklı dosya okuma şekilleri

import csv
with open("file.csv") as file:
    content = csv.reader(file)
    for row in content:
        print(row)

import json
with open("file.json") as file:
    content = json.load(file)

content.get("bişi..")

import xml.etree.ElementTree as ET
content = ET.parse("jile.xml")
root = content.getroot()
for child in root:
    print(child.tag, child.text)
    
import docx2txt
content = docx2txt.process("file.docx")
print(content)

from PyPDF2 import PdfFileReader
file = open("file.pdf", "rb")
content = PdfFileReader(file)
text = content.getPage(0)
text.extractText()

content.numPages













#%% Site engelleme

import time
from datetime import datetime as dt

hostsPath = r"C:\Windows\System32\drivers\etc\hosts"
redirect="127.0.0.1"
websites= ["www.youtube","youtube.com"]
while True:
    if dt(dt.now().year,dt.now().month,dt.now().day, 9) < dt.now() < dt(dt.now().year,dt.now().month,dt.now().day,18):
        print("Olmaz kardeş")
        with open(hostsPath,'r+') as file:
            content=file.read()
            for site in websites:
                if site in content:
                    pass
                else:
                    file.write(redirect+" "+site+"\n")
    else:
        with open(hostsPath,'r+') as file:
            content=file.readlines()
            file.seek(0)
            for line in content:
                if not any(site in line for site in websites):
                    file.write(line)
            file.truncate()
        print("Allowed access")
    time.sleep(5)
    
    












































































